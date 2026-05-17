from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook

from .types import SourceFormat


class SourceParser:
    def parse(self, source_format: SourceFormat, payload: str | bytes | list[dict[str, Any]] | dict[str, Any]) -> list[dict[str, Any]]:
        if isinstance(payload, list):
            return payload
        if isinstance(payload, dict):
            if source_format == SourceFormat.GEOJSON:
                return self._parse_geojson(payload)
            return [payload]
        match source_format:
            case SourceFormat.JSON:
                parsed = json.loads(payload)
                return parsed if isinstance(parsed, list) else [parsed]
            case SourceFormat.GEOJSON:
                return self._parse_geojson(json.loads(payload))
            case SourceFormat.CSV:
                text = payload.decode() if isinstance(payload, bytes) else payload
                return list(csv.DictReader(io.StringIO(text)))
            case SourceFormat.XML:
                text = payload.decode() if isinstance(payload, bytes) else payload
                root = ElementTree.fromstring(text)
                return [{el.tag: el.text for el in list(child)} for child in list(root)]
            case SourceFormat.EXCEL:
                return self._parse_excel(payload)
        raise ValueError(f'Unsupported source format: {source_format}')

    @staticmethod
    def _parse_geojson(payload: dict[str, Any]) -> list[dict[str, Any]]:
        return [{**f.get('properties', {}), 'geometry': f.get('geometry')} for f in payload.get('features', [])]

    @staticmethod
    def _parse_excel(payload: str | bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(filename=payload) if isinstance(payload, str) and Path(payload).exists() else load_workbook(filename=io.BytesIO(payload if isinstance(payload, bytes) else payload.encode()))
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]
        return [{str(header[i]): value for i, value in enumerate(row)} for row in sheet.iter_rows(min_row=2, values_only=True)]
